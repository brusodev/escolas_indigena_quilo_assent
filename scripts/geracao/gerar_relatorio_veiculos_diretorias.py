#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RELATÓRIO EXCEL - VEÍCULOS POR DIRETORIA
Análise específica das diretorias que atendem escolas indígenas, quilombolas e de assentamento
"""

import pandas as pd
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def carregar_dados():
    """Carrega dados das escolas e veículos"""
    print("📖 Carregando dados...")

    # Carregar dados das escolas
    arquivo_escolas = (
        "dados/excel/distancias_escolas_diretorias_completo_63_corrigido.xlsx"
    )
    if not os.path.exists(arquivo_escolas):
        print(f"❌ Arquivo de escolas não encontrado: {arquivo_escolas}")
        return None, None

    df_escolas = pd.read_excel(arquivo_escolas)
    print(f"✅ Escolas carregadas: {len(df_escolas)}")

    # Carregar dados dos veículos ORIGINAIS CORRETOS
    arquivo_veiculos = "dados/json/dados_veiculos_originais_corretos.json"
    if not os.path.exists(arquivo_veiculos):
        print(f"❌ Arquivo de veículos não encontrado: {arquivo_veiculos}")
        return df_escolas, None

    with open(arquivo_veiculos, "r", encoding="utf-8") as f:
        dados_veiculos_reais = json.load(f)

    # Calcular total de veículos
    total_veiculos = sum(
        diretoria["total"] for diretoria in dados_veiculos_reais.values()
    )
    print(
        f"✅ Dados de veículos ORIGINAIS CORRETOS carregados: {total_veiculos} veículos"
    )

    return df_escolas, dados_veiculos_reais


def analisar_diretorias_veiculos(df_escolas, dados_veiculos_reais):
    """Analisa a distribuição de veículos por diretoria usando dados ORIGINAIS CORRETOS"""
    print("🔍 Analisando distribuição de veículos ORIGINAIS CORRETOS...")

    # Agrupar escolas por diretoria
    diretorias_analise = []

    for diretoria in df_escolas["DE_Responsavel"].unique():
        escolas_diretoria = df_escolas[df_escolas["DE_Responsavel"] == diretoria]

        # Contar tipos de escola
        indigenas = len(
            escolas_diretoria[escolas_diretoria["Tipo_Escola"] == "Indígena"]
        )
        quilombolas = len(
            escolas_diretoria[
                escolas_diretoria["Tipo_Escola"] == "Quilombola/Assentamento"
            ]
        )
        total_escolas = len(escolas_diretoria)

        # Obter dados REAIS de veículos - procurar com e sem acentos/espaços
        total_veiculos = 0
        s1_veiculos = 0
        s2_veiculos = 0
        s2_4x4_veiculos = 0

        # Variações possíveis do nome da diretoria
        nomes_possiveis = [
            diretoria,
            diretoria.upper(),
            diretoria.strip(),
            diretoria.strip().upper(),
            diretoria.replace("Ã", "A").replace("Õ", "O").replace("Ç", "C"),
            diretoria.replace("ã", "a").replace("õ", "o").replace("ç", "c"),
        ]

        for nome in nomes_possiveis:
            if nome in dados_veiculos_reais:
                veiculos_info = dados_veiculos_reais[nome]
                total_veiculos = veiculos_info.get("total", 0)
                s1_veiculos = veiculos_info.get("s1", 0)
                s2_veiculos = veiculos_info.get("s2", 0)
                s2_4x4_veiculos = veiculos_info.get("s2_4x4", 0)
                break

        if total_veiculos == 0:
            print(f"⚠️  Diretoria sem dados de veículos: {diretoria}")

        # Calcular estatísticas de distância
        dist_media = escolas_diretoria["Distancia_KM"].mean()
        dist_maxima = escolas_diretoria["Distancia_KM"].max()
        dist_minima = escolas_diretoria["Distancia_KM"].min()

        # Contar escolas por zona
        rurais = len(escolas_diretoria[escolas_diretoria["Zona"] == "Rural"])
        urbanas = len(escolas_diretoria[escolas_diretoria["Zona"] == "Urbana"])

        # Classificação de prioridade baseada em distância e número de escolas
        if dist_media >= 60 and total_escolas >= 3:
            prioridade = "ALTA"
        elif dist_media >= 40 or total_escolas >= 5:
            prioridade = "MÉDIA"
        else:
            prioridade = "BAIXA"

        diretorias_analise.append(
            {
                "Diretoria": diretoria,
                "Total_Veiculos": total_veiculos,
                "Veiculos_S1": s1_veiculos,
                "Veiculos_S2": s2_veiculos,
                "Veiculos_S2_4x4": s2_4x4_veiculos,
                "Total_Escolas": total_escolas,
                "Escolas_Indigenas": indigenas,
                "Escolas_Quilombolas_Assentamento": quilombolas,
                "Escolas_Zona_Rural": rurais,
                "Escolas_Zona_Urbana": urbanas,
                "Distancia_Media_KM": round(dist_media, 2),
                "Distancia_Maxima_KM": round(dist_maxima, 2),
                "Distancia_Minima_KM": round(dist_minima, 2),
                "Veiculos_por_Escola": (
                    round(total_veiculos / total_escolas, 2) if total_escolas > 0 else 0
                ),
                "Prioridade_Logistica": prioridade,
                "Codigo_DE": (
                    escolas_diretoria["DE_Codigo"].iloc[0]
                    if len(escolas_diretoria) > 0
                    else ""
                ),
            }
        )

    # Converter para DataFrame e ordenar por número de veículos (maior para menor)
    df_diretorias = pd.DataFrame(diretorias_analise)
    df_diretorias = df_diretorias.sort_values("Total_Veiculos", ascending=False)

    print(f"✅ Análise concluída: {len(df_diretorias)} diretorias analisadas")

    return df_diretorias


def criar_relatorio_excel(df_diretorias, df_escolas, dados_veiculos_reais):
    """Cria o relatório Excel com múltiplas abas usando dados ORIGINAIS CORRETOS"""
    print("📊 Criando relatório Excel com dados ORIGINAIS CORRETOS...")

    # Calcular total de veículos dos dados originais corretos
    total_veiculos_reais = sum(
        diretoria["total"] for diretoria in dados_veiculos_reais.values()
    )

    # Criar workbook
    wb = Workbook()

    # Estilos
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2F5597", end_color="2F5597", fill_type="solid"
    )
    normal_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ===== ABA 1: RESUMO EXECUTIVO =====
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Executivo"

    # Título
    ws_resumo["A1"] = "RELATÓRIO DE VEÍCULOS POR DIRETORIA"
    ws_resumo["A1"].font = Font(name="Arial", size=16, bold=True, color="2F5597")
    ws_resumo.merge_cells("A1:H1")

    ws_resumo["A2"] = (
        "Sistema de Gestão de Escolas Indígenas, Quilombolas e Assentamentos"
    )
    ws_resumo["A2"].font = Font(name="Arial", size=12, italic=True)
    ws_resumo.merge_cells("A2:H2")

    ws_resumo["A3"] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_resumo["A3"].font = Font(name="Arial", size=10)
    ws_resumo.merge_cells("A3:H3")

    # Estatísticas gerais
    row = 5
    stats = [
        ("📊 ESTATÍSTICAS GERAIS", ""),
        ("Total de Diretorias:", len(df_diretorias)),
        ("Total de Veículos ORIGINAIS:", total_veiculos_reais),
        ("Total de Escolas:", len(df_escolas)),
        (
            "Escolas Indígenas:",
            len(df_escolas[df_escolas["Tipo_Escola"] == "Indígena"]),
        ),
        (
            "Escolas Quilombolas/Assentamento:",
            len(df_escolas[df_escolas["Tipo_Escola"] == "Quilombola/Assentamento"]),
        ),
        ("", ""),
        ("🚗 DISTRIBUIÇÃO DE VEÍCULOS ORIGINAIS", ""),
        (
            "Diretoria com mais veículos:",
            f"{df_diretorias.iloc[0]['Diretoria']} ({df_diretorias.iloc[0]['Total_Veiculos']} veículos)",
        ),
        (
            "Média de veículos por diretoria:",
            f"{df_diretorias['Total_Veiculos'].mean():.1f}",
        ),
        (
            "Média de escolas por diretoria:",
            f"{df_diretorias['Total_Escolas'].mean():.1f}",
        ),
        ("", ""),
        ("🎯 PRIORIDADES LOGÍSTICAS", ""),
        (
            "Diretorias ALTA prioridade:",
            len(df_diretorias[df_diretorias["Prioridade_Logistica"] == "ALTA"]),
        ),
        (
            "Diretorias MÉDIA prioridade:",
            len(df_diretorias[df_diretorias["Prioridade_Logistica"] == "MÉDIA"]),
        ),
        (
            "Diretorias BAIXA prioridade:",
            len(df_diretorias[df_diretorias["Prioridade_Logistica"] == "BAIXA"]),
        ),
        ("", ""),
        ("🚛 TIPOS DE VEÍCULOS", ""),
        ("Total S1 (pequenos):", df_diretorias["Veiculos_S1"].sum()),
        ("Total S2 (médios):", df_diretorias["Veiculos_S2"].sum()),
        ("Total S2 4x4 (terreno):", df_diretorias["Veiculos_S2_4x4"].sum()),
    ]

    for desc, valor in stats:
        ws_resumo[f"A{row}"] = desc
        ws_resumo[f"B{row}"] = valor
        if desc.startswith(("📊", "🚗", "🎯")):
            ws_resumo[f"A{row}"].font = Font(
                name="Arial", size=12, bold=True, color="2F5597"
            )
        else:
            ws_resumo[f"A{row}"].font = Font(name="Arial", size=10, bold=True)
        ws_resumo[f"B{row}"].font = normal_font
        row += 1

    # ===== ABA 2: VEÍCULOS POR DIRETORIA =====
    ws_veiculos = wb.create_sheet("Veiculos por Diretoria")

    # Cabeçalhos
    headers = [
        "Diretoria",
        "Total Veículos",
        "Veículos S1",
        "Veículos S2",
        "Veículos S2 4x4",
        "Total Escolas",
        "Escolas Indígenas",
        "Escolas Quilombolas/Assentamento",
        "Escolas Rurais",
        "Escolas Urbanas",
        "Distância Média (km)",
        "Distância Máxima (km)",
        "Veículos/Escola",
        "Prioridade Logística",
        "Código DE",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_veiculos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Dados
    for row_idx, (_, row_data) in enumerate(df_diretorias.iterrows(), 2):
        for col_idx, value in enumerate(
            [
                row_data["Diretoria"],
                row_data["Total_Veiculos"],
                row_data["Veiculos_S1"],
                row_data["Veiculos_S2"],
                row_data["Veiculos_S2_4x4"],
                row_data["Total_Escolas"],
                row_data["Escolas_Indigenas"],
                row_data["Escolas_Quilombolas_Assentamento"],
                row_data["Escolas_Zona_Rural"],
                row_data["Escolas_Zona_Urbana"],
                row_data["Distancia_Media_KM"],
                row_data["Distancia_Maxima_KM"],
                row_data["Veiculos_por_Escola"],
                row_data["Prioridade_Logistica"],
                row_data["Codigo_DE"],
            ],
            1,
        ):
            cell = ws_veiculos.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = (
                center_align
                if col_idx > 1
                else Alignment(horizontal="left", vertical="center")
            )

            # Colorir por prioridade
            if col_idx == 14:  # Coluna de prioridade (agora é a 14ª)
                if value == "ALTA":
                    cell.fill = PatternFill(
                        start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
                    )
                    cell.font = Font(name="Arial", size=10, bold=True, color="C62828")
                elif value == "MÉDIA":
                    cell.fill = PatternFill(
                        start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
                    )
                    cell.font = Font(name="Arial", size=10, bold=True, color="E65100")
                else:
                    cell.fill = PatternFill(
                        start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"
                    )
                    cell.font = Font(name="Arial", size=10, bold=True, color="2E7D32")

    # Ajustar larguras das colunas
    column_widths = [25, 12, 10, 10, 12, 12, 15, 25, 12, 12, 15, 15, 12, 15, 12]
    for col, width in enumerate(column_widths, 1):
        ws_veiculos.column_dimensions[
            ws_veiculos.cell(row=1, column=col).column_letter
        ].width = width

    # ===== ABA 3: APENAS ESCOLAS INDÍGENAS =====
    ws_indigenas = wb.create_sheet("Diretorias Indigenas")

    df_indigenas = df_diretorias[df_diretorias["Escolas_Indigenas"] > 0].copy()
    df_indigenas = df_indigenas.sort_values("Escolas_Indigenas", ascending=False)

    # Cabeçalhos específicos para indígenas
    headers_ind = [
        "Diretoria",
        "Veículos",
        "Escolas Indígenas",
        "Total Escolas",
        "% Indígenas",
        "Distância Média",
        "Prioridade",
    ]

    for col, header in enumerate(headers_ind, 1):
        cell = ws_indigenas.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(
            start_color="1B5E20", end_color="1B5E20", fill_type="solid"
        )
        cell.alignment = center_align
        cell.border = border

    # Dados das diretorias com escolas indígenas
    for row_idx, (_, row_data) in enumerate(df_indigenas.iterrows(), 2):
        perc_indigenas = (
            row_data["Escolas_Indigenas"] / row_data["Total_Escolas"]
        ) * 100

        for col_idx, value in enumerate(
            [
                row_data["Diretoria"],
                row_data["Total_Veiculos"],
                row_data["Escolas_Indigenas"],
                row_data["Total_Escolas"],
                f"{perc_indigenas:.1f}%",
                f"{row_data['Distancia_Media_KM']:.1f} km",
                row_data["Prioridade_Logistica"],
            ],
            1,
        ):
            cell = ws_indigenas.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = (
                center_align
                if col_idx > 1
                else Alignment(horizontal="left", vertical="center")
            )

    # ===== ABA 4: APENAS ESCOLAS QUILOMBOLAS/ASSENTAMENTO =====
    ws_quilombolas = wb.create_sheet("Diretorias Quilombolas")

    df_quilombolas = df_diretorias[
        df_diretorias["Escolas_Quilombolas_Assentamento"] > 0
    ].copy()
    df_quilombolas = df_quilombolas.sort_values(
        "Escolas_Quilombolas_Assentamento", ascending=False
    )

    # Cabeçalhos específicos para quilombolas
    headers_qui = [
        "Diretoria",
        "Veículos",
        "Escolas Quilombolas/Assentamento",
        "Total Escolas",
        "% Quilombolas",
        "Distância Média",
        "Prioridade",
    ]

    for col, header in enumerate(headers_qui, 1):
        cell = ws_quilombolas.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(
            start_color="4A148C", end_color="4A148C", fill_type="solid"
        )
        cell.alignment = center_align
        cell.border = border

    # Dados das diretorias com escolas quilombolas
    for row_idx, (_, row_data) in enumerate(df_quilombolas.iterrows(), 2):
        perc_quilombolas = (
            row_data["Escolas_Quilombolas_Assentamento"] / row_data["Total_Escolas"]
        ) * 100

        for col_idx, value in enumerate(
            [
                row_data["Diretoria"],
                row_data["Total_Veiculos"],
                row_data["Escolas_Quilombolas_Assentamento"],
                row_data["Total_Escolas"],
                f"{perc_quilombolas:.1f}%",
                f"{row_data['Distancia_Media_KM']:.1f} km",
                row_data["Prioridade_Logistica"],
            ],
            1,
        ):
            cell = ws_quilombolas.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = (
                center_align
                if col_idx > 1
                else Alignment(horizontal="left", vertical="center")
            )

    # Ajustar larguras para as abas específicas
    for ws in [ws_indigenas, ws_quilombolas]:
        column_widths_esp = [25, 10, 20, 12, 12, 15, 12]
        for col, width in enumerate(column_widths_esp, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Salvar arquivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"relatorios/excel/Relatorio_Veiculos_por_Diretoria_{timestamp}.xlsx"

    os.makedirs("relatorios/excel", exist_ok=True)
    wb.save(nome_arquivo)

    print(f"✅ Relatório Excel criado: {nome_arquivo}")

    return nome_arquivo, df_diretorias


def main():
    """Função principal"""
    print("🚗 GERADOR DE RELATÓRIO - VEÍCULOS POR DIRETORIA")
    print("=" * 70)
    print(f"📅 Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(
        "🎯 Foco: Diretorias que atendem escolas indígenas, quilombolas e assentamentos"
    )
    print()

    try:
        # Carregar dados
        df_escolas, dados_veiculos_reais = carregar_dados()
        if df_escolas is None or dados_veiculos_reais is None:
            print("❌ Falha ao carregar dados")
            return False

        # Analisar distribuição
        df_diretorias = analisar_diretorias_veiculos(df_escolas, dados_veiculos_reais)

        # Criar relatório
        nome_arquivo, df_resultado = criar_relatorio_excel(
            df_diretorias, df_escolas, dados_veiculos_reais
        )

        # Estatísticas finais
        print("\n📊 ESTATÍSTICAS DO RELATÓRIO:")
        print(f"   📋 Total de diretorias analisadas: {len(df_resultado)}")
        print(
            f"   🚗 Total de veículos distribuídos: {df_resultado['Total_Veiculos'].sum()}"
        )
        print(
            f"   🏛️ Diretorias com escolas indígenas: {len(df_resultado[df_resultado['Escolas_Indigenas'] > 0])}"
        )
        print(
            f"   🏡 Diretorias com escolas quilombolas/assentamento: {len(df_resultado[df_resultado['Escolas_Quilombolas_Assentamento'] > 0])}"
        )
        print(
            f"   🎯 Diretorias de ALTA prioridade: {len(df_resultado[df_resultado['Prioridade_Logistica'] == 'ALTA'])}"
        )
        print()

        # Top 5 diretorias com mais veículos
        print("🏆 TOP 5 DIRETORIAS COM MAIS VEÍCULOS:")
        top_5 = df_resultado.head(5)
        for i, (_, row) in enumerate(top_5.iterrows(), 1):
            print(
                f"   {i}. {row['Diretoria']}: {row['Total_Veiculos']} veículos "
                f"({row['Total_Escolas']} escolas)"
            )

        print(f"\n📁 Arquivo salvo: {nome_arquivo}")
        print("\n🎉 Relatório de veículos por diretoria gerado com sucesso!")

        return True

    except Exception as e:
        print(f"❌ Erro ao gerar relatório: {e}")
        import traceback

        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
