# -*- coding: utf-8 -*-
import os
import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import matplotlib.pyplot as plt
import seaborn as sns

# Configurar encoding UTF-8 para Windows
if sys.platform.startswith("win"):
    os.environ["PYTHONIOENCODING"] = "utf-8"


def carregar_dados_integrados():
    """Carrega e integra todos os dados: escolas, veículos e supervisão"""

    print("Carregando dados integrados...")

    # Dados existentes
    df_escolas = pd.read_excel("distancias_escolas_diretorias.xlsx")

    # Novos dados
    try:
        df_veiculos = pd.read_excel("QUANTIDADE DE VEÍCULOS LOCADOS - DIRETORIAS.xlsx")
        print(f"✅ Dados de veículos carregados: {len(df_veiculos)} diretorias")
    except FileNotFoundError:
        print("❌ Arquivo de veículos não encontrado. Criando dados de exemplo...")
        df_veiculos = pd.DataFrame()

    try:
        df_gep = pd.read_excel("GEP.xlsx")
        print(f"✅ Dados de supervisão carregados: {len(df_gep)} diretorias")
    except FileNotFoundError:
        print("❌ Arquivo GEP não encontrado. Criando dados de exemplo...")
        df_gep = pd.DataFrame()

    return df_escolas, df_veiculos, df_gep


def analisar_demanda_veiculos(df_escolas, df_veiculos, df_gep):
    """Analisa a demanda de veículos baseada na distância e distribuição das escolas"""

    print("\n🔍 Analisando demanda de veículos por diretoria...")

    # Agrupar escolas por diretoria
    analise_diretorias = (
        df_escolas.groupby("Nome_Diretoria")
        .agg(
            {
                "Nome_Escola": "count",
                "Distancia_KM": ["mean", "max", "min"],
                "Tipo_Escola": lambda x: ", ".join(x.unique()),
            }
        )
        .round(2)
    )

    # Achatar colunas multi-level
    analise_diretorias.columns = [
        "Total_Escolas",
        "Distancia_Media",
        "Distancia_Maxima",
        "Distancia_Minima",
        "Tipos_Escola",
    ]
    analise_diretorias = analise_diretorias.reset_index()

    # Calcular escolas distantes (>50km)
    escolas_distantes = (
        df_escolas[df_escolas["Distancia_KM"] > 50]
        .groupby("Nome_Diretoria")
        .size()
        .reset_index(name="Escolas_Distantes")
    )
    analise_diretorias = analise_diretorias.merge(
        escolas_distantes, on="Nome_Diretoria", how="left"
    )
    analise_diretorias["Escolas_Distantes"] = analise_diretorias[
        "Escolas_Distantes"
    ].fillna(0)

    # Padronizar nomes das diretorias para merge
    analise_diretorias["Diretoria_Padrao"] = analise_diretorias[
        "Nome_Diretoria"
    ].str.upper()
    df_veiculos["Diretoria_Padrao"] = df_veiculos["DIRETORIA"].str.upper()
    df_gep["Diretoria_Padrao"] = df_gep[
        "DIRETORIA DE ENSINO SOB SUPERVISÃO"
    ].str.upper()

    # Integrar dados de veículos
    if not df_veiculos.empty:
        analise_diretorias = analise_diretorias.merge(
            df_veiculos[
                [
                    "Diretoria_Padrao",
                    "QUANTIDADE S-1",
                    "QUANTIDADE S-2",
                    "QUANTIDADE S-2 4X4 ",
                ]
            ],
            on="Diretoria_Padrao",
            how="left",
        )
        # Calcular total de veículos
        analise_diretorias["Total_Veiculos"] = (
            analise_diretorias["QUANTIDADE S-1"].fillna(0)
            + analise_diretorias["QUANTIDADE S-2"].fillna(0)
            + analise_diretorias["QUANTIDADE S-2 4X4 "].fillna(0)
        )
    else:
        analise_diretorias["Total_Veiculos"] = 0

    # Integrar dados de supervisão
    if not df_gep.empty:
        analise_diretorias = analise_diretorias.merge(
            df_gep[
                ["Diretoria_Padrao", "REGIÃO", "SUPERVISOR GEP", "QUANTIDADE DE DEs"]
            ],
            on="Diretoria_Padrao",
            how="left",
        )

    # Calcular índice de necessidade de veículos
    analise_diretorias["Indice_Necessidade"] = (
        (
            analise_diretorias["Escolas_Distantes"] * 0.4
        )  # 40% peso para escolas distantes
        + (
            analise_diretorias["Distancia_Media"] / 20 * 0.3
        )  # 30% peso para distância média
        + (
            analise_diretorias["Total_Escolas"] / 5 * 0.3
        )  # 30% peso para total de escolas
    ).round(2)

    # Recomendação de veículos baseada no índice
    def recomendar_veiculos(row):
        indice = row["Indice_Necessidade"]
        veiculos_atuais = row["Total_Veiculos"]

        if indice >= 3:
            recomendacao = 3
        elif indice >= 2:
            recomendacao = 2
        elif indice >= 1:
            recomendacao = 1
        else:
            recomendacao = 0

        diferenca = recomendacao - veiculos_atuais
        return recomendacao, diferenca

    analise_diretorias[["Veiculos_Recomendados", "Diferenca_Veiculos"]] = (
        analise_diretorias.apply(recomendar_veiculos, axis=1, result_type="expand")
    )

    return analise_diretorias


def gerar_relatorio_frota_integrado(analise_diretorias):
    """Gera relatório Excel integrado com análise de frota"""

    print("\n📋 Gerando relatório integrado de frota...")

    wb = Workbook()

    # Estilos
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2C3E50", end_color="2C3E50", fill_type="solid"
    )
    data_font = Font(name="Calibri", size=10)
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Aba 1: Análise Integrada
    ws_principal = wb.active
    ws_principal.title = "Análise Integrada"

    # Título
    ws_principal.merge_cells("A1:P1")
    ws_principal["A1"] = "ANÁLISE INTEGRADA: ESCOLAS, FROTA E SUPERVISÃO"
    ws_principal["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws_principal["A1"].fill = PatternFill(
        start_color="E74C3C", end_color="E74C3C", fill_type="solid"
    )
    ws_principal["A1"].alignment = center_alignment

    # Headers
    headers = [
        "Diretoria",
        "Total Escolas",
        "Escolas Distantes (>50km)",
        "Distância Média (km)",
        "Distância Máxima (km)",
        "Veículos S-1",
        "Veículos S-2",
        "Veículos S-2 4x4",
        "Total Veículos",
        "Índice Necessidade",
        "Veículos Recomendados",
        "Diferença",
        "Região",
        "Supervisor GEP",
        "Status",
        "Prioridade",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws_principal.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Dados
    for row_idx, (_, row) in enumerate(analise_diretorias.iterrows(), start=4):
        # Determinar status e prioridade
        if row["Diferenca_Veiculos"] > 0:
            status = "REQUER VEÍCULOS"
            prioridade = "ALTA" if row["Diferenca_Veiculos"] >= 2 else "MÉDIA"
        elif row["Diferenca_Veiculos"] < 0:
            status = "EXCESSO DE VEÍCULOS"
            prioridade = "BAIXA"
        else:
            status = "ADEQUADO"
            prioridade = "BAIXA"

        dados_row = [
            row["Nome_Diretoria"],
            int(row["Total_Escolas"]),
            int(row["Escolas_Distantes"]),
            row["Distancia_Media"],
            row["Distancia_Maxima"],
            int(
                row.get("QUANTIDADE S-1", 0)
                if pd.notna(row.get("QUANTIDADE S-1", 0))
                else 0
            ),
            int(
                row.get("QUANTIDADE S-2", 0)
                if pd.notna(row.get("QUANTIDADE S-2", 0))
                else 0
            ),
            int(
                row.get("QUANTIDADE S-2 4X4 ", 0)
                if pd.notna(row.get("QUANTIDADE S-2 4X4 ", 0))
                else 0
            ),
            int(row["Total_Veiculos"]),
            row["Indice_Necessidade"],
            int(row["Veiculos_Recomendados"]),
            int(row["Diferenca_Veiculos"]),
            row.get("REGIÃO", "N/A"),
            row.get("SUPERVISOR GEP", "N/A"),
            status,
            prioridade,
        ]

        for col, value in enumerate(dados_row, start=1):
            cell = ws_principal.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.border = border

            # Colorir baseado na prioridade
            if col == 16:  # Coluna prioridade
                if value == "ALTA":
                    cell.fill = PatternFill(
                        start_color="FADBD8", end_color="FADBD8", fill_type="solid"
                    )
                elif value == "MÉDIA":
                    cell.fill = PatternFill(
                        start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"
                    )
                else:
                    cell.fill = PatternFill(
                        start_color="D5F4E6", end_color="D5F4E6", fill_type="solid"
                    )

    # Ajustar larguras
    column_widths = [25, 12, 20, 15, 15, 12, 12, 15, 12, 15, 18, 12, 15, 20, 18, 12]
    for i, width in enumerate(column_widths, start=1):
        from openpyxl.utils import get_column_letter

        ws_principal.column_dimensions[get_column_letter(i)].width = width

    # Aba 2: Resumo Executivo Frota
    ws_resumo = wb.create_sheet("Resumo Frota")

    ws_resumo.merge_cells("A1:D1")
    ws_resumo["A1"] = "RESUMO EXECUTIVO - ANÁLISE DE FROTA"
    ws_resumo["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws_resumo["A1"].alignment = center_alignment

    # Estatísticas
    total_escolas = analise_diretorias["Total_Escolas"].sum()
    total_escolas_distantes = analise_diretorias["Escolas_Distantes"].sum()
    total_veiculos_atual = analise_diretorias["Total_Veiculos"].sum()
    total_veiculos_recomendado = analise_diretorias["Veiculos_Recomendados"].sum()
    diretorias_precisam_veiculos = len(
        analise_diretorias[analise_diretorias["Diferenca_Veiculos"] > 0]
    )

    stats = [
        ("Total de Escolas Atendidas:", total_escolas),
        ("Escolas Distantes (>50km):", total_escolas_distantes),
        ("% Escolas Distantes:", f"{(total_escolas_distantes/total_escolas*100):.1f}%"),
        ("Total de Veículos Atual:", total_veiculos_atual),
        ("Total de Veículos Recomendado:", total_veiculos_recomendado),
        (
            "Diferença de Veículos Necessária:",
            total_veiculos_recomendado - total_veiculos_atual,
        ),
        ("Diretorias que Precisam de Veículos:", diretorias_precisam_veiculos),
    ]

    for i, (label, value) in enumerate(stats, start=3):
        ws_resumo[f"A{i}"] = label
        ws_resumo[f"B{i}"] = value
        ws_resumo[f"A{i}"].font = Font(bold=True)

    # Salvar arquivo
    arquivo_nome = "Analise_Integrada_Escolas_Frota_Supervisao.xlsx"
    wb.save(arquivo_nome)

    print(f"✅ Relatório integrado salvo: {arquivo_nome}")
    return arquivo_nome


def main():
    """Função principal para análise integrada"""

    print("=" * 80)
    print("🚗📊 ANÁLISE INTEGRADA: ESCOLAS × FROTA × SUPERVISÃO")
    print("=" * 80)

    # Carregar dados
    df_escolas, df_veiculos, df_gep = carregar_dados_integrados()

    # Analisar demanda
    analise_diretorias = analisar_demanda_veiculos(df_escolas, df_veiculos, df_gep)

    # Mostrar diretorias que mais precisam de veículos
    print("\n🔥 TOP 10 DIRETORIAS QUE MAIS PRECISAM DE VEÍCULOS:")
    print("-" * 80)
    top_demanda = analise_diretorias.nlargest(10, "Diferenca_Veiculos")
    for _, row in top_demanda.iterrows():
        if row["Diferenca_Veiculos"] > 0:
            print(f"📍 {row['Nome_Diretoria']}")
            print(
                f"   Escolas distantes: {int(row['Escolas_Distantes'])} | Veículos atuais: {int(row['Total_Veiculos'])} | Recomendado: {int(row['Veiculos_Recomendados'])} | Falta: +{int(row['Diferenca_Veiculos'])}"
            )

    # Gerar relatório
    arquivo = gerar_relatorio_frota_integrado(analise_diretorias)

    print(f"\n✅ Análise concluída!")
    print(f"📁 Relatório salvo em: {arquivo}")

    return analise_diretorias


if __name__ == "__main__":
    main()
